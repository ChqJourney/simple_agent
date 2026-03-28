from __future__ import annotations

import re
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
except Exception:  # pragma: no cover - exercised via runtime validation
    load_workbook = None
    get_column_letter = None


MAX_EXCEL_FILE_BYTES = 15 * 1024 * 1024
MAX_TITLE_CHARS = 80


def _require_openpyxl() -> None:
    if load_workbook is None or get_column_letter is None:
        raise RuntimeError("openpyxl is not installed. Install `openpyxl` to use Excel document tools.")


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return " ".join(str(value).split()).strip()


def _make_anchor(text: str) -> str:
    anchor = re.sub(r"[^\w\s-]", "", text.lower()).strip()
    anchor = re.sub(r"[\s_]+", "-", anchor)
    return anchor


def _truncate_title(text: str) -> str:
    normalized = " ".join(str(text or "").split()).strip()
    if len(normalized) <= MAX_TITLE_CHARS:
        return normalized
    return normalized[:MAX_TITLE_CHARS].rstrip() + "..."


def _format_row_text(cells: list[dict[str, Any]]) -> str:
    parts = [str(cell.get("text") or "").strip() for cell in cells if str(cell.get("text") or "").strip()]
    return " | ".join(parts)


class ExcelReader:
    def __init__(self, xlsx_path: str | Path):
        _require_openpyxl()
        self.xlsx_path = Path(xlsx_path)
        if not self.xlsx_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.xlsx_path}")
        file_size = self.xlsx_path.stat().st_size
        if file_size > MAX_EXCEL_FILE_BYTES:
            raise ValueError(f"File too large: {file_size} bytes (max: {MAX_EXCEL_FILE_BYTES} bytes)")

        self.workbook = load_workbook(self.xlsx_path, read_only=True, data_only=True)
        self.sheets: list[dict[str, Any]] = [self._serialize_sheet(ws) for ws in self.workbook.worksheets]

    def _serialize_sheet(self, worksheet: Any) -> dict[str, Any]:
        max_row = int(getattr(worksheet, "max_row", 0) or 0)
        max_column = int(getattr(worksheet, "max_column", 0) or 0)
        rows: list[dict[str, Any]] = []
        non_empty_row_indexes: list[int] = []
        non_empty_column_indexes: set[int] = set()

        if max_row > 0 and max_column > 0:
            for row_index, row in enumerate(
                worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_column, values_only=True),
                start=1,
            ):
                cells: list[dict[str, Any]] = []
                has_text = False
                for column_index, value in enumerate(row, start=1):
                    text = _normalize_text(value)
                    if text:
                        has_text = True
                        non_empty_column_indexes.add(column_index)
                    cells.append(
                        {
                            "column_index": column_index,
                            "column_letter": get_column_letter(column_index),
                            "text": text,
                        }
                    )
                if has_text:
                    non_empty_row_indexes.append(row_index)
                rows.append(
                    {
                        "row_index": row_index,
                        "cells": cells,
                        "text": _format_row_text(cells),
                    }
                )

        if non_empty_row_indexes:
            row_start = min(non_empty_row_indexes)
            row_end = max(non_empty_row_indexes)
        else:
            row_start = 1
            row_end = 0

        if non_empty_column_indexes:
            column_start = min(non_empty_column_indexes)
            column_end = max(non_empty_column_indexes)
        else:
            column_start = 1
            column_end = 0

        trimmed_rows = []
        if row_end >= row_start and column_end >= column_start:
            for row in rows[row_start - 1:row_end]:
                trimmed_cells = [
                    cell for cell in row["cells"] if column_start <= int(cell["column_index"]) <= column_end
                ]
                trimmed_rows.append(
                    {
                        "row_index": row["row_index"],
                        "cells": trimmed_cells,
                        "text": _format_row_text(trimmed_cells),
                    }
                )

        header_row_index, column_headers = self._infer_header_row(trimmed_rows)
        return {
            "sheet_name": str(getattr(worksheet, "title", "Sheet")),
            "row_start": row_start,
            "row_end": row_end,
            "column_start": column_start,
            "column_end": column_end,
            "row_count": max(0, row_end - row_start + 1),
            "column_count": max(0, column_end - column_start + 1),
            "header_row_index": header_row_index,
            "column_headers": column_headers,
            "rows": trimmed_rows,
        }

    def _infer_header_row(self, rows: list[dict[str, Any]]) -> tuple[int | None, list[str]]:
        for row in rows[:5]:
            values = [str(cell.get("text") or "").strip() for cell in row["cells"] if str(cell.get("text") or "").strip()]
            if len(values) >= 2 and sum(len(value) for value in values) <= 200:
                return int(row["row_index"]), values
        return None, []

    @property
    def sheet_count(self) -> int:
        return len(self.sheets)

    def _get_sheet(self, sheet_name: str) -> dict[str, Any]:
        normalized = sheet_name.strip()
        for sheet in self.sheets:
            if sheet["sheet_name"] == normalized:
                return sheet
        raise ValueError(f"Sheet not found: {sheet_name}")

    def get_structure(self, max_nodes: int = 200) -> dict[str, Any]:
        nodes = []
        for sheet in self.sheets[:max_nodes]:
            title = str(sheet["sheet_name"])
            nodes.append(
                {
                    "title": title,
                    "level": 1,
                    "anchor": _make_anchor(title),
                    "locator": {
                        "sheet_name": title,
                        "row_start": int(sheet["row_start"]),
                        "row_end": int(sheet["row_end"]),
                        "column_start": int(sheet["column_start"]),
                        "column_end": int(sheet["column_end"]),
                    },
                    "sheet_summary": {
                        "row_count": int(sheet["row_count"]),
                        "column_count": int(sheet["column_count"]),
                        "header_row_index": sheet["header_row_index"],
                    },
                }
            )

        return {
            "xlsx_path": str(self.xlsx_path),
            "sheet_count": self.sheet_count,
            "structure_type": "excel_workbook_map",
            "items": nodes,
        }

    def search(
        self,
        query: str,
        *,
        mode: str = "plain",
        case_sensitive: bool = False,
        max_results: int = 50,
        context_rows: int = 2,
    ) -> dict[str, Any]:
        flags = 0 if case_sensitive else re.IGNORECASE
        pattern = re.compile(query if mode == "regex" else re.escape(query), flags)
        items: list[dict[str, Any]] = []

        for sheet in self.sheets:
            rows = list(sheet["rows"])
            non_empty_rows = [row for row in rows if row["text"]]
            column_headers = list(sheet.get("column_headers") or [])
            for row_position, row in enumerate(non_empty_rows):
                for cell in row["cells"]:
                    if len(items) >= max_results:
                        break
                    cell_text = str(cell.get("text") or "")
                    if not cell_text:
                        continue
                    match = pattern.search(cell_text)
                    if not match:
                        continue

                    start = max(0, row_position - context_rows)
                    end = min(len(non_empty_rows), row_position + context_rows + 1)
                    context_before = "\n".join(entry["text"] for entry in non_empty_rows[start:row_position] if entry["text"])
                    context_after = "\n".join(
                        entry["text"] for entry in non_empty_rows[row_position + 1:end] if entry["text"]
                    )

                    column_index = int(cell["column_index"])
                    column_header = ""
                    if column_headers:
                        relative_index = column_index - int(sheet["column_start"])
                        if 0 <= relative_index < len(column_headers):
                            column_header = str(column_headers[relative_index] or "")

                    items.append(
                        {
                            "source_type": "cell",
                            "sheet_name": str(sheet["sheet_name"]),
                            "row_index": int(row["row_index"]),
                            "column_index": column_index,
                            "column_letter": str(cell["column_letter"]),
                            "cell_ref": f"{cell['column_letter']}{row['row_index']}",
                            "column_header": column_header,
                            "row_text": str(row["text"]),
                            "match_text": match.group(0),
                            "text": cell_text,
                            "context_before": context_before,
                            "context_after": context_after,
                        }
                    )
                if len(items) >= max_results:
                    break
            if len(items) >= max_results:
                break

        return {
            "xlsx_path": str(self.xlsx_path),
            "sheet_count": self.sheet_count,
            "items": items,
        }

    def read_range(
        self,
        sheet_name: str,
        *,
        row_start: int | None = None,
        row_end: int | None = None,
        column_start: int | None = None,
        column_end: int | None = None,
        include_context: int = 0,
    ) -> dict[str, Any]:
        sheet = self._get_sheet(sheet_name)
        total_rows = int(sheet["row_end"])
        total_columns = int(sheet["column_end"])
        if total_rows == 0 or total_columns == 0:
            return {
                "xlsx_path": str(self.xlsx_path),
                "sheet_name": sheet_name,
                "row_start": 1,
                "row_end": 0,
                "column_start": 1,
                "column_end": 0,
                "header_row_index": None,
                "column_headers": [],
                "items": [],
            }

        normalized_row_start = int(row_start) if row_start is not None else int(sheet["row_start"])
        normalized_row_end = int(row_end) if row_end is not None else total_rows
        normalized_column_start = int(column_start) if column_start is not None else int(sheet["column_start"])
        normalized_column_end = int(column_end) if column_end is not None else total_columns

        if normalized_row_start < 1 or normalized_row_end < normalized_row_start:
            raise ValueError("Invalid row range: row_end must be >= row_start and row_start must be >= 1")
        if normalized_row_end > total_rows:
            raise ValueError(f"Row range {normalized_row_start}-{normalized_row_end} exceeds row count {total_rows}")
        if normalized_column_start < 1 or normalized_column_end < normalized_column_start:
            raise ValueError("Invalid column range: column_end must be >= column_start and column_start must be >= 1")
        if normalized_column_end > total_columns:
            raise ValueError(
                f"Column range {normalized_column_start}-{normalized_column_end} exceeds column count {total_columns}"
            )

        requested = set(range(normalized_row_start, normalized_row_end + 1))
        selected = set(requested)
        if include_context > 0:
            selected.update(range(max(1, normalized_row_start - include_context), min(total_rows, normalized_row_end + include_context) + 1))

        header_row_index = int(sheet["header_row_index"]) if sheet.get("header_row_index") else None
        if header_row_index is not None:
            selected.add(header_row_index)

        items = []
        for row in sheet["rows"]:
            row_index = int(row["row_index"])
            if row_index not in selected:
                continue
            filtered_cells = [
                cell for cell in row["cells"] if normalized_column_start <= int(cell["column_index"]) <= normalized_column_end
            ]
            items.append(
                {
                    "row_index": row_index,
                    "requested": row_index in requested,
                    "is_header": row_index == header_row_index,
                    "cells": filtered_cells,
                    "text": _format_row_text(filtered_cells),
                }
            )

        return {
            "xlsx_path": str(self.xlsx_path),
            "sheet_name": sheet_name,
            "row_start": normalized_row_start,
            "row_end": normalized_row_end,
            "column_start": normalized_column_start,
            "column_end": normalized_column_end,
            "header_row_index": header_row_index,
            "column_headers": list(sheet.get("column_headers") or []),
            "items": items,
        }


def get_excel_structure(xlsx_path: str | Path, max_nodes: int = 200) -> dict[str, Any]:
    reader = ExcelReader(xlsx_path)
    return reader.get_structure(max_nodes=max_nodes)


def search_excel_workbook(
    xlsx_path: str | Path,
    query: str,
    *,
    mode: str = "plain",
    case_sensitive: bool = False,
    max_results: int = 50,
    context_rows: int = 2,
) -> dict[str, Any]:
    reader = ExcelReader(xlsx_path)
    return reader.search(
        query,
        mode=mode,
        case_sensitive=case_sensitive,
        max_results=max_results,
        context_rows=context_rows,
    )


def read_excel_range(
    xlsx_path: str | Path,
    sheet_name: str,
    *,
    row_start: int | None = None,
    row_end: int | None = None,
    column_start: int | None = None,
    column_end: int | None = None,
    include_context: int = 0,
) -> dict[str, Any]:
    reader = ExcelReader(xlsx_path)
    return reader.read_range(
        sheet_name,
        row_start=row_start,
        row_end=row_end,
        column_start=column_start,
        column_end=column_end,
        include_context=include_context,
    )
